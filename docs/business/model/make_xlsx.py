#!/usr/bin/env python3
"""Generate fedora-3yr-model.xlsx from build_model.py's assumptions.

The workbook is fully formula-driven: edit any blue input cell on the
Assumptions sheet and every scenario recalculates. Sheets:
  Summary      headline outcomes per scenario (formulas into scenario sheets)
  Assumptions  every input, low/medium/high columns, with source notes
  Low/Medium/High   36-month P&L grids mirroring build_model.run_scenario()

Run:  python3 docs/business/model/make_xlsx.py
Then recalculate with LibreOffice (see xlsx skill scripts/recalc.py) so
formula caches are populated.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_model as bm

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MONEY = "£#,##0;(£#,##0);-"
MONEY2 = "£#,##0.00;(£#,##0.00);-"
PCT = "0.0%"
NUM = "#,##0"
NUM2 = "0.00"

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=13, bold=True)
INPUT = Font(name="Arial", size=10, color="0000FF")
KEY_FILL = PatternFill("solid", fgColor="FFFF00")
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")

SCEN_COL = {"low": "B", "medium": "C", "high": "D"}
N_MONTHS = bm.MONTHS
FIRST_C = 2                       # months start in column B
LAST_L = get_column_letter(FIRST_C + N_MONTHS - 1)   # AK

# ---------------------------------------------------------------------------
# Assumption sheet layout definitions
# ---------------------------------------------------------------------------

# Structural constants promoted to visible inputs
STRUCTURAL = [
    ("n_cities", "Cities at launch", NUM, bm.N_CITIES,
     "Launch brief: one hunt per big city in 10 cities"),
    ("team_size", "Avg players per team (one purchase covers a team)", NUM2,
     bm.TEAM_SIZE, "Teams of 3-4; marketing-cac.md §5"),
    ("wave2_month", "Wave-2 month (paid catalogue + sponsors + ads)", NUM, bm.WAVE2_MONTH,
     "Launch brief: monetisation at second launch"),
    ("mau_lapse", "Lapsed players active per month (MAU factor)", PCT, 0.08,
     "Episodic use; D30 retention 3-5%, marketing-cac.md §2"),
    ("ota_new_share", "OTA players new to app (share)", PCT, 0.5,
     "OTA buyers partly overlap app installs"),
    ("home_share", "Home-city share of catalogue", PCT, 0.15,
     "1 of 10 launch cities is home; grows slightly"),
    ("build_discount", "Wave-2+ build discount vs first remote hunt", PCT, 0.9,
     "Repeat visits amortise city research"),
]

SHARED_DEFS = [
    ("hunt_build_home", "Hunt build — home city (£)", MONEY,
     "Contracted design+field+red-team+QA; competitive-response.md §3"),
    ("hunt_build_remote", "Hunt build — remote city (£)", MONEY,
     "+travel/local contractor (Blue Badge £353/day; trip ~$1,771)"),
    ("hunt_maintain_yr_home", "Maintenance/yr — home hunt (£)", MONEY,
     "Quarterly re-verification; world drifts"),
    ("hunt_maintain_yr_remote", "Maintenance/yr — remote hunt (£)", MONEY,
     "Local contractor passes; Field Agent-style spot checks"),
    ("redteam_compute_per_hunt", "AI red-team compute per hunt (£)", MONEY,
     "Frontier-model attack runs pre-ship (methodology step 2)"),
    ("platform_monthly", "Platform/tools per month (£)", MONEY,
     "Supabase + hosting + tooling at scale"),
    ("app_build_oneoff", "Native app build one-off (£)", MONEY,
     "Store-presence wrapper + polish; web app already exists"),
    ("fee_rate", "Fees on direct consumer revenue", PCT,
     "Web Stripe 1.5-2.9% blended with some 15% IAP; marketing-cac.md §4"),
    ("ota_commission", "OTA commission", PCT,
     "GetYourGuide/Viator take 20-30%; marketing-cac.md §6"),
    ("corp_delivery_rate", "Corporate delivery cost (% of corp revenue)", PCT,
     "Host on the day, props, admin"),
    ("wl_delivery_rate", "White-label delivery cost (% of WL revenue)", PCT,
     "Bespoke build + verification labour"),
    ("payer_churn_mo", "Repeat-buyer pool monthly churn", PCT,
     "Lapse rate of paying teams"),
]

SCEN_DEFS = [
    ("cpi", None, "Blended CPI (£/install)", MONEY2,
     "Tier-1 cities $1.50-6.00; marketing-cac.md §1", True),
    ("organic_mult_yr", 0, "Organic multiplier Y1", NUM2,
     "PR/creators/referral uplift; decays (Randonautica)", False),
    ("organic_mult_yr", 1, "Organic multiplier Y2", NUM2, "", False),
    ("organic_mult_yr", 2, "Organic multiplier Y3", NUM2, "", False),
    ("install_to_active", None, "Install → starts free hunt", PCT,
     "D1-retention-like; marketing-cac.md §2", True),
    ("backlog_conv", None, "Backlog engaged-team conversion", PCT,
     "Pre-wave-2 base converting when catalogue lands", False),
    ("ongoing_conv", None, "Engaged team → paid conversion", PCT,
     "Freemium 2.1% of installs ≈ 7-10% of activated; team-level", True),
    ("repeats_per_payer_yr", None, "Repeat purchases/payer/yr", NUM2,
     "Capped by catalogue releases per city", False),
    ("price_per_team", None, "Price per team (£, gross)", MONEY,
     "≈£11-14/head; Questo €8-20pp, CityDays £15-25pp", True),
    ("ota_teams_city_mo_yr", 0, "OTA teams/city/month Y1", NUM, "GYG/Viator listings from wave 2", False),
    ("ota_teams_city_mo_yr", 1, "OTA teams/city/month Y2", NUM, "", False),
    ("ota_teams_city_mo_yr", 2, "OTA teams/city/month Y3", NUM, "", False),
    ("mkt_launch_burst", None, "Launch burst, months 0-2 (£)", MONEY,
     "10-city campaign £250k/£650k/£1.5M yr1; marketing-cac.md §3", True),
    ("mkt_monthly_yr", 0, "Marketing/month Y1 after burst (£)", MONEY, "", False),
    ("mkt_monthly_yr", 1, "Marketing/month Y2 (£)", MONEY, "", False),
    ("mkt_monthly_yr", 2, "Marketing/month Y3 (£)", MONEY, "", False),
    ("sponsor_per_city_yr", None, "Sponsorship per live city/yr (£)", MONEY,
     "£2.5-4k / £8-15k / £30-60k; sponsorship-economics.md", True),
    ("sponsor_cities_ramp", 0, "Cities with sponsors Y1", NUM, "", False),
    ("sponsor_cities_ramp", 1, "Cities with sponsors Y2", NUM, "", False),
    ("sponsor_cities_ramp", 2, "Cities with sponsors Y3", NUM, "", False),
    ("wl_deals_yr", 0, "White-label deals Y1", NUM,
     "Councils/BIDs/museums; Ipswich £22.8k anchor", False),
    ("wl_deals_yr", 1, "White-label deals Y2", NUM, "", False),
    ("wl_deals_yr", 2, "White-label deals Y3", NUM, "", False),
    ("wl_value", None, "White-label deal value (£)", MONEY,
     "£2.5-5k / £10-25k / £50k+; sponsorship-economics.md §3", False),
    ("ad_rev_per_1k_mau_mo", None, "Ads £ per 1,000 MAU/month", MONEY,
     "$10-100; episodic use; sponsorship-economics.md §4", False),
    ("corp_events_mo_yr", 0, "Corporate events/month Y1", NUM2, "", False),
    ("corp_events_mo_yr", 1, "Corporate events/month Y2", NUM2, "", False),
    ("corp_events_mo_yr", 2, "Corporate events/month Y3", NUM2, "", False),
    ("corp_rev_per_event", None, "Revenue per corporate event (£)", MONEY,
     "ClueGo £55-65pp × 25-40 heads; competitor-pricing.md", True),
    ("staff_monthly_yr", 0, "Staff/month Y1 (£, loaded)", MONEY, "", False),
    ("staff_monthly_yr", 1, "Staff/month Y2 (£)", MONEY, "", False),
    ("staff_monthly_yr", 2, "Staff/month Y3 (£)", MONEY, "", False),
    ("new_hunts_yr", 0, "New paid hunts built Y1", NUM, "", False),
    ("new_hunts_yr", 1, "New paid hunts built Y2", NUM, "", False),
    ("new_hunts_yr", 2, "New paid hunts built Y3", NUM, "", False),
    ("prune_factor_yr", 0, "Catalogue maintained Y1", PCT, "Low prunes dead cities", False),
    ("prune_factor_yr", 1, "Catalogue maintained Y2", PCT, "", False),
    ("prune_factor_yr", 2, "Catalogue maintained Y3", PCT, "", False),
    ("competitor_entry_month", None, "Competitor entry month", NUM,
     "6-18mo after visible traction; competitive-response.md §5", True),
    ("post_entry_growth_dampener", None, "Post-entry growth dampener", PCT,
     "Escape-room convergence precedent", False),
    ("post_entry_price_factor", None, "Post-entry price factor", PCT,
     "Discount pressure after entry", False),
]

PCT_KEYS = {"fee_rate", "ota_commission", "corp_delivery_rate",
            "wl_delivery_rate", "payer_churn_mo", "install_to_active",
            "backlog_conv", "ongoing_conv", "prune_factor_yr",
            "post_entry_growth_dampener", "post_entry_price_factor",
            "mau_lapse", "ota_new_share", "home_share", "build_discount"}


def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Fedora — global launch model: assumptions"
    ws["A1"].font = TITLE
    ws["A2"] = ("All money GBP. Blue = input cells (edit these); yellow = key levers. "
                "Sources: docs/business/research/*.md")
    ws["A2"].font = ARIAL

    refs = {}          # ("shared"|scenario, key, idx) -> cell row
    r = 4
    ws.cell(r, 1, "STRUCTURE & SHARED").font = BOLD
    ws.cell(r, 1).fill = HDR_FILL
    r += 1
    for key, label, fmt, val, src in STRUCTURAL:
        ws.cell(r, 1, label).font = ARIAL
        c = ws.cell(r, 2, val)
        c.font = INPUT
        c.number_format = fmt
        ws.cell(r, 5, src).font = ARIAL
        refs[("shared", key)] = r
        r += 1
    for key, label, fmt, src in SHARED_DEFS:
        ws.cell(r, 1, label).font = ARIAL
        c = ws.cell(r, 2, bm.SHARED[key])
        c.font = INPUT
        c.number_format = fmt
        ws.cell(r, 5, src).font = ARIAL
        refs[("shared", key)] = r
        r += 1

    r += 1
    ws.cell(r, 1, "SCENARIO").font = BOLD
    ws.cell(r, 1).fill = HDR_FILL
    for i, name in enumerate(("Low", "Medium", "High")):
        ws.cell(r, 2 + i, name).font = BOLD
        ws.cell(r, 2 + i).fill = HDR_FILL
    ws.cell(r, 5, "Source / basis").font = BOLD
    r += 1
    for key, idx, label, fmt, src, is_key in SCEN_DEFS:
        ws.cell(r, 1, label).font = ARIAL
        for j, scen in enumerate(("low", "medium", "high")):
            v = bm.ASSUMPTIONS[scen][key]
            if idx is not None:
                v = v[idx]
            c = ws.cell(r, 2 + j, v)
            c.font = INPUT
            c.number_format = fmt
            if is_key:
                c.fill = KEY_FILL
        ws.cell(r, 5, src).font = ARIAL
        refs[(key, idx)] = r
        r += 1

    ws.column_dimensions["A"].width = 44
    for col in "BCD":
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["E"].width = 62
    return refs


ROWS = [
    (1, "Month", NUM), (2, "Year", NUM), (3, "Post-entry flag", NUM),
    (4, "Growth dampener", NUM2), (5, "Price per team (£)", MONEY2),
    (6, "Marketing spend (£)", MONEY), (7, "Installs", NUM),
    (8, "New active players", NUM), (9, "New active teams", NUM),
    (10, "Backlog pool (teams)", NUM), (11, "Backlog tranche", NUM),
    (12, "New paying teams", NUM), (13, "Repeat rate (monthly)", "0.000"),
    (14, "Repeat purchases", NUM), (15, "Paying-team base", NUM),
    (16, "Paid team purchases", NUM), (17, "Direct consumer net (£)", MONEY),
    (18, "OTA teams", NUM), (19, "OTA net (£)", MONEY),
    (20, "Cumulative players", NUM), (21, "MAU", NUM),
    (22, "Advertising (£)", MONEY), (23, "Sponsorship (£)", MONEY),
    (24, "White-label (£)", MONEY), (25, "Corporate revenue (£)", MONEY),
    (26, "TOTAL REVENUE (£)", MONEY), (27, "New hunts built", NUM2),
    (28, "Live hunts", NUM2), (29, "Content build (£)", MONEY),
    (30, "Content maintenance (£)", MONEY), (31, "Staff (£)", MONEY),
    (32, "Other costs (£)", MONEY), (33, "TOTAL COSTS (£)", MONEY),
    (34, "NET (£)", MONEY), (35, "Cumulative net (£)", MONEY),
    (36, "Cumulative revenue (£)", MONEY),
]


def build_scenario(wb, scen, refs):
    ws = wb.create_sheet(scen.capitalize())
    col = SCEN_COL[scen]

    def A(key):
        return f"Assumptions!$B${refs[('shared', key)]}"

    def S(key, idx=None):
        return f"Assumptions!${col}${refs[(key, idx)]}"

    def yr3(key):
        rr = [refs[(key, i)] for i in range(3)]
        return (f"CHOOSE({{Y}},Assumptions!${col}${rr[0]},"
                f"Assumptions!${col}${rr[1]},Assumptions!${col}${rr[2]})")

    for rnum, label, fmt in ROWS:
        c = ws.cell(rnum, 1, label)
        c.font = BOLD if rnum in (26, 33, 34, 35) else ARIAL

    for j in range(N_MONTHS):
        L = get_column_letter(FIRST_C + j)
        P = get_column_letter(FIRST_C + j - 1)
        first = j == 0
        Y = f"{L}2"

        def put(rnum, formula_or_val):
            c = ws.cell(rnum, FIRST_C + j)
            c.value = formula_or_val
            c.font = ARIAL
            c.number_format = ROWS[rnum - 1][2]

        put(1, j)
        put(2, f"=1+INT({L}1/12)")
        put(3, f"=IF({L}1>={S('competitor_entry_month')},1,0)")
        put(4, f"=IF({L}3=1,{S('post_entry_growth_dampener')},1)")
        put(5, f"={S('price_per_team')}*IF({L}3=1,{S('post_entry_price_factor')},1)")
        put(6, f"=IF({L}1<3,{S('mkt_launch_burst')}/3,"
               + yr3("mkt_monthly_yr").replace("{Y}", Y) + ")")
        put(7, f"={L}6/{S('cpi')}*" + yr3("organic_mult_yr").replace("{Y}", Y) + f"*{L}4")
        put(8, f"={L}7*{S('install_to_active')}")
        put(9, f"={L}8/{A('team_size')}")
        if first:
            put(10, f"={L}9")
            put(11, "=0")
        else:
            put(10, f"=IF({L}1<{A('wave2_month')},{P}10+{L}9,"
                    f"IF({L}1<={A('wave2_month')}+1,{P}10*2/3,0))")
            put(11, f"=IF({L}1<{A('wave2_month')},0,"
                    f"IF({L}1<={A('wave2_month')}+1,{P}10/3,"
                    f"IF({L}1={A('wave2_month')}+2,{P}10,0)))")
        put(12, f"={L}11*{S('backlog_conv')}"
                f"+IF({L}1>={A('wave2_month')},{L}9*{S('ongoing_conv')},0)")
        put(13, f"=MIN({S('repeats_per_payer_yr')},"
                + yr3("new_hunts_yr").replace("{Y}", Y)
                + f"/{A('n_cities')})/12")
        put(14, "=0" if first else f"={P}15*{L}13")
        put(15, f"={L}12" if first
            else f"={P}15*(1-{A('payer_churn_mo')})+{L}12")
        put(16, f"={L}12+{L}14")
        put(17, f"={L}16*{L}5*(1-{A('fee_rate')})")
        put(18, f"=IF({L}1>={A('wave2_month')},"
                + yr3("ota_teams_city_mo_yr").replace("{Y}", Y)
                + f"*{A('n_cities')}*{L}4,0)")
        put(19, f"={L}18*{L}5*(1-{A('ota_commission')})")
        base20 = f"{L}8+{L}18*{A('team_size')}*{A('ota_new_share')}"
        put(20, f"={base20}" if first else f"={P}20+{base20}")
        put(21, f"={L}8+{A('mau_lapse')}*{L}20")
        put(22, f"=IF({L}1>={A('wave2_month')},{L}21/1000*{S('ad_rev_per_1k_mau_mo')},0)")
        put(23, f"=IF({L}1>={A('wave2_month')},"
                + yr3("sponsor_cities_ramp").replace("{Y}", Y)
                + f"*{S('sponsor_per_city_yr')}/12,0)")
        put(24, "=" + yr3("wl_deals_yr").replace("{Y}", Y) + f"*{S('wl_value')}/12")
        put(25, "=" + yr3("corp_events_mo_yr").replace("{Y}", Y)
                + f"*{S('corp_rev_per_event')}")
        put(26, f"={L}17+{L}19+{L}22+{L}23+{L}24+{L}25")
        put(27, "=" + yr3("new_hunts_yr").replace("{Y}", Y) + "/12")
        put(28, f"={A('n_cities')}+{L}27" if first else f"={P}28+{L}27")
        put(29, f"=IF({L}1=0,{A('hunt_build_home')}+{A('redteam_compute_per_hunt')}"
                f"+({A('n_cities')}-1)*({A('hunt_build_remote')}+{A('redteam_compute_per_hunt')}),0)"
                f"+{L}27*({A('hunt_build_remote')}*{A('build_discount')}+{A('redteam_compute_per_hunt')})")
        put(30, f"={L}28*" + yr3("prune_factor_yr").replace("{Y}", Y)
                + f"*({A('home_share')}*{A('hunt_maintain_yr_home')}"
                f"+(1-{A('home_share')})*{A('hunt_maintain_yr_remote')})/12")
        put(31, "=" + yr3("staff_monthly_yr").replace("{Y}", Y))
        put(32, f"=IF({L}1=0,{A('app_build_oneoff')},0)+{A('platform_monthly')}"
                f"+{L}25*{A('corp_delivery_rate')}+{L}24*{A('wl_delivery_rate')}")
        put(33, f"={L}6+{L}29+{L}30+{L}31+{L}32")
        put(34, f"={L}26-{L}33")
        put(35, f"={L}34" if first else f"={P}35+{L}34")
        put(36, f"={L}26" if first else f"={P}36+{L}26")

    ws.column_dimensions["A"].width = 26
    for j in range(N_MONTHS):
        ws.column_dimensions[get_column_letter(FIRST_C + j)].width = 11
    ws.freeze_panes = "B1"
    return ws


def build_summary(wb, refs):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Fedora — 3-year global-launch outcomes"
    ws["A1"].font = TITLE
    ws["A2"] = ("10 cities at launch, first hunt free, wave-2 monetisation at month 9. "
                "All values recalculate from the Assumptions sheet.")
    ws["A2"].font = ARIAL

    GREEN = Font(name="Arial", size=10, color="008000")
    r = 4
    ws.cell(r, 1, "").font = BOLD
    for i, name in enumerate(("Low", "Medium", "High")):
        c = ws.cell(r, 2 + i, name)
        c.font = BOLD
        c.fill = HDR_FILL

    def row(label, per_scen_formula, fmt=MONEY, bold=False):
        nonlocal r
        r += 1
        c = ws.cell(r, 1, label)
        c.font = BOLD if bold else ARIAL
        for i, scen in enumerate(("Low", "Medium", "High")):
            col = SCEN_COL[scen.lower()]
            f = per_scen_formula.format(S=scen, C=col)
            cc = ws.cell(r, 2 + i, f)
            cc.font = GREEN
            cc.number_format = fmt

    for y in (1, 2, 3):
        row(f"Year {y} revenue (£)",
            f"=SUMIF({{S}}!$B$2:${LAST_L}$2,{y},{{S}}!$B$26:${LAST_L}$26)")
    for y in (1, 2, 3):
        row(f"Year {y} costs (£)",
            f"=SUMIF({{S}}!$B$2:${LAST_L}$2,{y},{{S}}!$B$33:${LAST_L}$33)")
    for y in (1, 2, 3):
        row(f"Year {y} net (£)",
            f"=SUMIF({{S}}!$B$2:${LAST_L}$2,{y},{{S}}!$B$34:${LAST_L}$34)")
    row("Cumulative net, 36 months (£)", f"={{S}}!${LAST_L}$35", bold=True)
    row("Cash trough (£)", f"=MIN({{S}}!$B$35:${LAST_L}$35)", bold=True)
    row("Players reached (app)", f"=SUM({{S}}!$B$8:${LAST_L}$8)", NUM, bold=True)

    r += 1
    ws.cell(r + 0, 1, "")
    row("Competitor entry month",
        "=Assumptions!${C}$%d" % refs[("competitor_entry_month", None)], NUM)
    entry = "MATCH(Assumptions!${C}$%d,{S}!$B$1:$%s$1,0)-1" % (
        refs[("competitor_entry_month", None)], LAST_L)
    row("Revenue banked at entry (£)",
        f"=INDEX({{S}}!$B$36:${LAST_L}$36,{entry})", bold=True)
    row("Cash position at entry (£)",
        f"=INDEX({{S}}!$B$35:${LAST_L}$35,{entry})")
    row("Players reached at entry",
        f"=INDEX({{S}}!$B$20:${LAST_L}$20,{entry})", NUM)

    r += 1
    labels = [("Direct consumer", 17), ("OTA", 19), ("Advertising", 22),
              ("Sponsorship", 23), ("White-label", 24), ("Corporate", 25)]
    for lab, rn in labels:
        row(f"Year 3 {lab} (£)",
            f"=SUMIF({{S}}!$B$2:${LAST_L}$2,3,{{S}}!$B${rn}:${LAST_L}${rn})")

    ws.column_dimensions["A"].width = 34
    for col in "BCD":
        ws.column_dimensions[col].width = 15


def main():
    wb = Workbook()
    wb.remove(wb.active)
    # Assumptions first so cell refs exist; Summary inserted at index 0
    refs = build_assumptions(wb)
    for scen in ("low", "medium", "high"):
        build_scenario(wb, scen, refs)
    build_summary(wb, refs)
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "fedora-3yr-model.xlsx")
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
